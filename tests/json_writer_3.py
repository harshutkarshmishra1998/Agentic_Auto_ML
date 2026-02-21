"""
json_writer_3.py

Convert JSONL file in project_root/data/ to human readable Excel workbook.
Exports only the last N records (user controlled).
"""

import json
from pathlib import Path
from openpyxl import Workbook


JSONL_FILENAME = "preprocesses_1.jsonl"
OUTPUT_FILENAME = "preprocesses_1.xlsx"


# -----------------------------------------------------
# helpers
# -----------------------------------------------------
def _project_root():
    return Path(__file__).resolve().parents[1]


def _fmt(v):
    if v is None:
        return ""
    if isinstance(v, (list, dict)):
        return json.dumps(v)
    return v


def _write_dict_sheet(wb, sheet_name, data: dict):
    ws = wb.create_sheet(sheet_name[:31])
    ws.append(["key", "value"])
    for k, v in data.items():
        ws.append([k, _fmt(v)])


def _write_list_of_dicts(wb, sheet_name, rows):
    ws = wb.create_sheet(sheet_name[:31])
    if not rows:
        return

    headers = sorted({k for r in rows for k in r.keys()})
    ws.append(headers)

    for r in rows:
        ws.append([_fmt(r.get(h)) for h in headers])


def _write_simple_list(wb, sheet_name, values):
    ws = wb.create_sheet(sheet_name[:31])
    ws.append(["value"])
    for v in values:
        ws.append([_fmt(v)])


# -----------------------------------------------------
# robust JSON loader
# -----------------------------------------------------
def _load_jsonl(path: Path):
    text = path.read_text(encoding="utf-8")

    objs = []
    buffer = ""
    depth = 0
    in_string = False
    escape = False

    for ch in text:
        buffer += ch

        if ch == '"' and not escape:
            in_string = not in_string

        if ch == "\\" and not escape:
            escape = True
            continue
        escape = False

        if in_string:
            continue

        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                objs.append(json.loads(buffer))
                buffer = ""

    return objs


# -----------------------------------------------------
# exporter
# -----------------------------------------------------
def export_jsonl_to_excel(jsonl_path: Path, output_xlsx: Path, n: int | None = None):

    records = _load_jsonl(jsonl_path)

    if not records:
        print("No records found")
        return

    # -------- last N selection --------
    if n is not None:
        records = records[-n:]

    wb = Workbook()
    wb.remove(wb.active)

    for i, rec in enumerate(records, start=1):
        prefix = f"record_{i}_"

        flat = {}
        nested = {}

        for k, v in rec.items():
            if isinstance(v, (dict, list)):
                nested[k] = v
            else:
                flat[k] = v

        _write_dict_sheet(wb, prefix + "metadata", flat)

        for key, value in nested.items():
            sheet = prefix + key

            if isinstance(value, list):
                if value and isinstance(value[0], dict):
                    _write_list_of_dicts(wb, sheet, value)
                else:
                    _write_simple_list(wb, sheet, value)
            else:
                _write_dict_sheet(wb, sheet, value)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)

    print(f"\nExcel written → {output_xlsx}")
    print(f"Records exported → {len(records)}")


# -----------------------------------------------------
# entry
# -----------------------------------------------------
if __name__ == "__main__":

    root = _project_root()
    data_dir = root / "data"

    jsonl_file = data_dir / JSONL_FILENAME
    output_file = data_dir / OUTPUT_FILENAME

    if not jsonl_file.exists():
        raise FileNotFoundError(jsonl_file)

    export_jsonl_to_excel(jsonl_file, output_file, 1)