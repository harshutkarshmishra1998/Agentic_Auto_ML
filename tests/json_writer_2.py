# import json
# import csv
# from pathlib import Path


# COLUMN_FIELDS = [
#     "column_name",
#     "technical_type",
#     "semantic_type",
#     "role",
#     "cardinality_level",
#     "missing_pct",
#     "missing_pattern",
#     "distribution_shape",
#     "outliers_present",
#     "unique_ratio",
#     "encoding_required",
#     "time_dependent",
#     "unit_scale",
#     "text_complexity",
#     "category_imbalance",
#     "correlation_strength",
#     "transform_hint",
#     "modeling_hint",
#     "data_quality_flags",
#     "is_constant",
# ]


# def _fmt(v):
#     if v is None:
#         return ""
#     if isinstance(v, float):
#         return round(v, 6)
#     if isinstance(v, list):
#         return ", ".join(map(str, v))
#     if isinstance(v, dict):
#         return str(v)
#     return v


# def export_column_profiles_jsonl_to_csv(jsonl_path: str, csv_output: str, last_n: int):
#     path = Path(jsonl_path)
#     if not path.exists():
#         raise FileNotFoundError(path)

#     with path.open("r", encoding="utf-8") as f:
#         lines = f.readlines()

#     start = max(0, len(lines) - last_n)

#     rows_to_write = []

#     for idx, line in enumerate(lines[start:], start=start + 1):
#         record = json.loads(line)

#         dataset_name = record.get("dataset_file_name")
#         dataset_path = record.get("dataset_file_path")
#         column_profiles = record.get("column_profiles", [])

#         for col in column_profiles:
#             row = {
#                 "record_index": idx,
#                 "dataset_file_name": dataset_name,
#                 "dataset_file_path": dataset_path,
#             }

#             for field in COLUMN_FIELDS:
#                 row[field] = _fmt(col.get(field))

#             rows_to_write.append(row)

#     # write CSV
#     out = Path(csv_output)
#     out.parent.mkdir(parents=True, exist_ok=True)

#     with out.open("w", newline="", encoding="utf-8") as f:
#         writer = csv.DictWriter(
#             f,
#             fieldnames=[
#                 "record_index",
#                 "dataset_file_name",
#                 "dataset_file_path",
#                 *COLUMN_FIELDS,
#             ],
#         )
#         writer.writeheader()
#         writer.writerows(rows_to_write)

#     print(f"CSV written → {out}")


# # ---------------------------------------------------
# # run
# # ---------------------------------------------------
# if __name__ == "__main__":

#     JSONL_FILE = "data/column_inspection.jsonl"
#     CSV_FILE = "data/json_writer_2.csv"
#     LAST_N = 1

#     export_column_profiles_jsonl_to_csv(JSONL_FILE, CSV_FILE, LAST_N)

# inspection_bundle_to_xlsx.py

import json
from pathlib import Path
from openpyxl import Workbook


# --------------------------------------------------
# helpers
# --------------------------------------------------
def _fmt(v):
    if v is None:
        return ""
    if isinstance(v, (list, dict)):
        return json.dumps(v)
    return v


def _write_table(ws, rows, headers):
    ws.append(headers)
    for r in rows:
        ws.append([_fmt(r.get(h)) for h in headers])


def _write_simple_list(ws, values, header="value"):
    ws.append([header])
    for v in values:
        ws.append([_fmt(v)])


def _write_dependency_graph(ws, graph: dict):
    ws.append(["feature", "depends_on"])
    for k, deps in graph.items():
        if not deps:
            ws.append([k, ""])
        else:
            for d in deps:
                ws.append([k, d])


# --------------------------------------------------
# main exporter
# --------------------------------------------------
def export_full_inspection_bundle(jsonl_path: str, n: int, output_xlsx: str):
    path = Path(jsonl_path)
    if not path.exists():
        raise FileNotFoundError(path)

    lines = path.read_text(encoding="utf-8").splitlines()
    if not lines:
        print("JSONL empty")
        return

    start = max(0, len(lines) - n)

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    for record_idx, line in enumerate(lines[start:], start=start + 1):
        obj = json.loads(line)

        dataset_name = obj.get("dataset_file_name")
        dataset_path = obj.get("dataset_file_path")

        prefix = f"rec_{record_idx}_"

        # ---------------- metadata ----------------
        ws = wb.create_sheet(prefix + "metadata")
        ws.append(["dataset_file_name", dataset_name])
        ws.append(["dataset_file_path", dataset_path])

        # ---------------- column_profiles ----------------
        column_profiles = obj.get("column_profiles", [])
        if column_profiles:
            headers = list(column_profiles[0].keys())
            ws = wb.create_sheet(prefix + "column_profiles")
            _write_table(ws, column_profiles, headers)

        # ---------------- correlation_pairs ----------------
        corr = obj.get("correlation_pairs", [])
        if corr:
            headers = list(corr[0].keys())
            ws = wb.create_sheet(prefix + "correlation_pairs")
            _write_table(ws, corr, headers)

        # ---------------- redundant_features ----------------
        redundant = obj.get("redundant_features", [])
        if redundant:
            headers = list(redundant[0].keys())
            ws = wb.create_sheet(prefix + "redundant_features")
            _write_table(ws, redundant, headers)

        # ---------------- derived_relationships ----------------
        derived = obj.get("derived_relationships", [])
        if derived:
            headers = list(derived[0].keys())
            ws = wb.create_sheet(prefix + "derived_relationships")
            _write_table(ws, derived, headers)

        # ---------------- dependency_graph ----------------
        dep_graph = obj.get("dependency_graph", {})
        if dep_graph:
            ws = wb.create_sheet(prefix + "dependency_graph")
            _write_dependency_graph(ws, dep_graph)

        # ---------------- drop_recommendations ----------------
        drops = obj.get("drop_recommendations", [])
        if drops:
            ws = wb.create_sheet(prefix + "drop_recommendations")
            _write_simple_list(ws, drops, "feature_to_drop")

    output = Path(output_xlsx)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    print(f"Excel bundle exported → {output}")


# --------------------------------------------------
# run
# --------------------------------------------------
if __name__ == "__main__":

    JSONL_FILE = "data/column_inspection.jsonl"
    OUTPUT_XLSX = "data/inspection_bundle.xlsx"
    n = 1

    export_full_inspection_bundle(JSONL_FILE, n, OUTPUT_XLSX)